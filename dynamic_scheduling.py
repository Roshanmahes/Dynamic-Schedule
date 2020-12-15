import numpy as np
import pandas as pd
from openpyxl import load_workbook
import urllib
from io import BytesIO

def load_workbook_from_url(url):
    file = urllib.request.urlopen(url).read()
    return load_workbook(filename=BytesIO(file))

def dynamic_schedule(mean, SCV, omega, n, i, k, u):
    """
    Computes the table of optimal interarrival times tau_{i}(k,u).
    """
    
    Delta = 0.01
    m = int(u / Delta)

    # retrieve files
    url = 'https://github.com/Roshanmahes/Dynamic-Schedule/blob/main/output/'
    file_name = f'-SCV-{round(SCV,2)}-omega-{round(omega,1)}-n-20-m-250'
    file_name_tau = url + 'tau' + file_name.replace('.', '_') + '.xlsx?raw=true'
    file_name_xi = url + 'xi' + file_name.replace('.', '_') + '.xlsx?raw=true'

    work_sheets_tau = load_workbook_from_url(file_name_tau).worksheets
    work_sheets_xi = load_workbook_from_url(file_name_xi).worksheets

    # retrieve cost
    df_xi_sheet = pd.DataFrame(work_sheets_xi[i-n-1].values)
    cost = round(df_xi_sheet[k-1][m] * mean,2)

    # create table
    df_list = []

    for i in range(1,n):

        df_sheet = pd.DataFrame(work_sheets_tau[i-n].values)
        df_row = df_sheet.iloc[m,:i] * mean
        df_list += [[f'{df_row[i]:.2f}' for i in range(len(df_row))]]

    df = pd.DataFrame(df_list, columns=range(1,n)).fillna('')
    df.index = range(1,n)

    df.loc[n,:] = range(1,n)
    df.index = list(range(1,n)) + ['i / k']

    df.index.name = 'i'
    df.reset_index(level=0, inplace=True)

    return df, cost


def style_table(n, i, k):
    """
    Styles the table of the interarrival times.
    """

    style_data_conditional = [
        {   # remove animation when selected
            'if': {'state': 'selected'},
            'background-color': 'inherit !important',
            'border': 'inherit !important',
        },
        {   # row header
            'if': {'column_id': 'i'},
            'background-color': '#FAFAFA',
            'fontWeight': 'bold',
        },
        {   # column header
            'if': {'row_index': n-1},
            'background-color': '#FAFAFA',
            'fontWeight': 'bold',
        },
    ] + [
        {   # border if i == k
            'if': {'row_index': i, 'column_id': i+1},
        } for i in range(n)
    ] + [
        {   # no border if i > k
            'if': {'row_index': i, 'column_id': i+j},
            'border': '1px solid #ffffff',
        } for i in range(n) for j in range(2,n)
    ] + [
        {   # border if i == k
            'if': {'row_index': i, 'column_id': i+1},
            'border': '1px solid #d5d5d5',
        } for i in range(n)
    ] + [
        {   # highlight row
            'if': {'row_index': i-1, 'column_id': l},
            'background-color': '#e0f1f0',
        } for l in ['i'] + list(range(k))
    ] + [
        {   # highlight column
            'if': {'row_index': l, 'column_id': k},
            'background-color': '#e0f1f0',
        } for l in range(i,n)
    ] + [
        {   # interarrival time tau_i(k,u)
            'if': {'row_index': i-1, 'column_id': k},
            'background-color': '#8ed1ca',
            'border': '1px solid #028073',
        },
        {   # row index i
            'if': {'row_index': i-1, 'column_id': 'i'},
            'background-color': '#B7E1DD',
            'border': '1px solid #bbb',
        },
        {   # column index k
            'if': {'row_index': n-1, 'column_id': k},
            'background-color': '#B7E1DD',
            'border': '1px solid #bbb',
        },
    ]

    return style_data_conditional
